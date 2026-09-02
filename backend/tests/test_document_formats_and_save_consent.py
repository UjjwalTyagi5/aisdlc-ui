"""Documents in every format the user asked for, and stored only when they say so.

TWO THINGS.

1. FORMATS. Word, Excel and slides had generators; PDF had none at all — the only
   mention of .pdf in the agents was reading uploads. weasyprint is a dependency but
   does not import on this platform (missing GTK/Pango native libraries), so the
   renderer is reportlab, which does.

2. CONSENT. Generated files were persisted to the project's artifacts, and uploaded to
   Blob, the moment a tool produced them. Downloading a file in chat and publishing it
   to shared, durable, project-wide storage are different acts with different
   consequences, and the second one is the user's call. Files are now staged until
   they agree.
"""
from __future__ import annotations

import os
import sys
from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


# -- PDF rendering -------------------------------------------------------------


@pytest.mark.unit
def test_a_pdf_is_actually_written(tmp_path):
    from shared.tools.pdf_render import markdown_to_pdf

    out = str(tmp_path / "doc.pdf")
    markdown_to_pdf("# Title\n\nSome prose.", out, title="Doc")
    assert os.path.getsize(out) > 0
    with open(out, "rb") as fh:
        assert fh.read(5) == b"%PDF-"


@pytest.mark.unit
def test_markdown_structure_survives(tmp_path):
    """Headings, lists, a table and a fenced block must not throw. They exercise every
    branch of the parser, which is where a renderer usually breaks."""
    from shared.tools.pdf_render import markdown_to_pdf

    md = (
        "# H1\n## H2\n\ntext\n\n- a\n- b\n\n1. one\n2. two\n\n"
        "| Field | Value |\n|---|---|\n| Owner | Ana |\n\n"
        "```\ncode()\n```\n\n---\n[link](https://example.com)\n"
    )
    out = str(tmp_path / "rich.pdf")
    markdown_to_pdf(md, out)
    assert os.path.getsize(out) > 0


@pytest.mark.unit
def test_characters_that_break_naive_markup_are_escaped(tmp_path):
    """'<' and '&' are markup to reportlab's paragraph parser. A requirement saying
    "latency < 300ms" or "R&D" would otherwise fail the whole document."""
    from shared.tools.pdf_render import markdown_to_pdf

    out = str(tmp_path / "esc.pdf")
    markdown_to_pdf("Latency < 300ms for R&D <tags> & more", out)
    assert os.path.getsize(out) > 0


@pytest.mark.unit
def test_an_empty_document_still_produces_a_file(tmp_path):
    """Returning success with no file is how an agent tells a user their document is
    ready when it is not."""
    from shared.tools.pdf_render import markdown_to_pdf

    out = str(tmp_path / "empty.pdf")
    markdown_to_pdf("", out)
    assert os.path.getsize(out) > 0


# -- every format the user asked for has a tool --------------------------------


@pytest.mark.unit
@pytest.mark.parametrize(
    ("tool_name", "fmt"),
    [
        ("markdowntodoc", "Word"),
        ("markdowntopdf", "PDF"),
        ("generate_planning_sheet", "Excel"),
        ("generate_ppt", "PowerPoint"),
    ],
)
def test_requirements_can_produce_each_format(tool_name, fmt):
    from agents_orchestrator.requirements_agent.agents import planning

    assert tool_name in {t.name for t in planning.tools}, f"no {fmt} generator"


@pytest.mark.unit
@pytest.mark.parametrize("tool_name", ["save_architecture", "save_architecture_pdf"])
def test_design_can_produce_word_and_pdf(tool_name):
    from agents_orchestrator.design_architecture_agent.agents import architecture

    assert tool_name in {t.name for t in architecture.tools}


@pytest.mark.unit
@pytest.mark.parametrize(
    ("ext", "ctype"),
    [
        (".docx", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
        (".pdf", "application/pdf"),
        (".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        (".pptx", "application/vnd.openxmlformats-officedocument.presentationml.presentation"),
    ],
)
def test_each_format_stores_with_the_right_content_type(ext, ctype):
    """A wrong content type is served as application/octet-stream, so the browser
    downloads a file it could have displayed."""
    from shared.services.chat_artifacts import _CONTENT_TYPES

    assert _CONTENT_TYPES.get(ext) == ctype


# -- consent before anything is stored -----------------------------------------


@pytest.fixture(autouse=True)
def _clean_pending():
    from shared.services import chat_artifacts

    chat_artifacts._PENDING.clear()
    yield
    chat_artifacts._PENDING.clear()


def _ctx(consented: bool, session="s1"):
    import config.ws_helper as ws
    from shared.services import chat_artifacts as ca

    return (
        patch.object(ca, "get_tenant_id", lambda: "t1"),
        patch.object(ca, "get_project_id", lambda: "p1"),
        patch.object(ca, "get_session_id", lambda: session),
        patch.object(ws, "get_consequential_approved", lambda: consented),
    )


@pytest.mark.unit
async def test_without_consent_nothing_is_stored(tmp_path):
    """THE HEADLINE. No Artifact row, no Blob upload — the file is staged instead."""
    from shared.services import chat_artifacts as ca

    f = tmp_path / "brd.docx"
    f.write_bytes(b"x")
    store = AsyncMock()
    a, b, c, d = _ctx(consented=False)
    with a, b, c, d, patch("shared.services.artifact_store.store_artifact", store):
        await ca.register_generated_file("brd.docx", str(f), "http://x/brd.docx", stage="requirements")

    store.assert_not_awaited()
    pending = ca.pending_for_session("s1")
    assert [p["filename"] for p in pending] == ["brd.docx"]


@pytest.mark.unit
async def test_regenerating_the_same_file_does_not_queue_it_twice(tmp_path):
    """The user would otherwise be asked about one document two or three times."""
    from shared.services import chat_artifacts as ca

    f = tmp_path / "brd.docx"
    f.write_bytes(b"x")
    a, b, c, d = _ctx(consented=False)
    with a, b, c, d:
        for _ in range(3):
            await ca.register_generated_file("brd.docx", str(f), "u", stage="requirements")
    assert len(ca.pending_for_session("s1")) == 1


@pytest.mark.unit
async def test_saving_pending_files_stores_them(tmp_path):
    from shared.services import chat_artifacts as ca

    f = tmp_path / "brd.docx"
    f.write_bytes(b"x")
    a, b, c, d = _ctx(consented=False)
    with a, b, c, d:
        await ca.register_generated_file("brd.docx", str(f), "u", stage="requirements")

    called = {}

    async def _fake(filename, file_path, url, *, stage, consented=None):
        called["consented"] = consented
        called["filename"] = filename

    with patch.object(ca, "register_generated_file", _fake):
        stored, failed, not_uploaded = await ca.save_pending_artifacts("s1")

    assert stored == ["brd.docx"] and failed == []
    # Reaching save IS the consent — re-reading the per-turn flag here would refuse the
    # very action the user just granted.
    assert called["consented"] is True
    assert ca.pending_for_session("s1") == []


@pytest.mark.unit
async def test_saving_with_nothing_pending_is_not_an_error():
    from shared.services import chat_artifacts as ca

    assert await ca.save_pending_artifacts("nobody") == ([], [], [])


@pytest.mark.unit
async def test_consent_in_the_same_turn_stores_immediately(tmp_path):
    """"generate the BRD and save it" must not need a second round trip."""
    from shared.services import chat_artifacts as ca

    f = tmp_path / "brd.docx"
    f.write_bytes(b"x")
    a, b, c, d = _ctx(consented=True)
    with a, b, c, d, patch.object(ca, "_get_or_create_chat_run", AsyncMock(return_value=None)):
        await ca.register_generated_file("brd.docx", str(f), "u", stage="requirements")
    # Staged nothing: it took the store path (which stopped at the run lookup here).
    assert ca.pending_for_session("s1") == []


# -- the agents are told to ask ------------------------------------------------


@pytest.mark.unit
@pytest.mark.parametrize(
    ("module", "attr"),
    [
        ("agents_orchestrator.requirements_agent.agents.planning", "SYS_MESSAGE"),
        ("agents_orchestrator.design_architecture_agent.agents.architecture", "DESIGN_SYS_MESSAGE"),
    ],
)
def test_both_prompts_require_asking_before_saving(module, attr):
    import importlib

    prompt = getattr(importlib.import_module(module), attr, None)
    if prompt is None:  # requirements exports INGESTION_SYS_MESSAGE
        prompt = getattr(importlib.import_module(module), "INGESTION_SYS_MESSAGE")
    flat = " ".join(prompt.split())
    assert "SAVING DOCUMENTS TO THE PROJECT (ASK FIRST)" in flat
    assert "ONLY after they say yes" in flat
    assert "never claim a document was added to the project" in flat


@pytest.mark.unit
@pytest.mark.parametrize(
    "module",
    [
        "agents_orchestrator.requirements_agent.agents.planning",
        "agents_orchestrator.design_architecture_agent.agents.architecture",
    ],
)
def test_both_agents_expose_the_explicit_save(module):
    import importlib

    mod = importlib.import_module(module)
    assert "save_to_project_artifacts" in {t.name for t in mod.tools}


# -- both agents can produce every format --------------------------------------


@pytest.mark.unit
@pytest.mark.parametrize(
    "module",
    [
        "agents_orchestrator.requirements_agent.agents.planning",
        "agents_orchestrator.design_architecture_agent.agents.architecture",
    ],
)
def test_both_agents_export_every_document_format(module):
    """Design had no Excel generator at all, and its Word/PDF tools were hard-wired to
    the architecture document — it could not export arbitrary content in any format."""
    import importlib

    mod = importlib.import_module(module)
    names = {t.name for t in mod.tools}
    assert "export_document" in names            # word / pdf / excel / markdown
    assert "generate_diagram" in names           # image
    assert "generate_ppt" in names               # slides


@pytest.mark.unit
@pytest.mark.parametrize("ext", [".docx", ".pdf", ".xlsx", ".md"])
async def test_the_dispatcher_writes_each_format(tmp_path, ext):
    from shared.tools.doc_export import render_document

    md = "# Title\n\ntext\n\n| A | B |\n|---|---|\n| 1 | 2 |\n"
    out = str(tmp_path / f"doc{ext}")
    await render_document(md, out, title="Doc")
    assert os.path.getsize(out) > 0


@pytest.mark.unit
async def test_an_unsupported_format_is_refused_by_name(tmp_path):
    """Silently writing a .docx for a .rtf request is how a user ends up with a file
    they cannot open."""
    from shared.tools.doc_export import render_document

    with pytest.raises(ValueError) as e:
        await render_document("x", str(tmp_path / "doc.rtf"))
    assert ".rtf" in str(e.value)


@pytest.mark.unit
@pytest.mark.parametrize(
    ("given", "expected"),
    [
        ("report.pdf", "report.pdf"),
        ("report", "report.docx"),               # extension from the default
        ("../../etc/passwd.pdf", "passwd.pdf"),  # path separators stripped
        ("", "d.docx"),
    ],
)
def test_the_model_cannot_choose_a_path(given, expected):
    """The FILENAME comes from the model. A separator in it would write outside the
    session's output directory."""
    from shared.tools.doc_export import normalise_filename

    assert normalise_filename(given, "d.docx") == expected


# -- excel exports tables ------------------------------------------------------


@pytest.mark.unit
def test_excel_makes_one_sheet_per_table_named_from_the_heading(tmp_path):
    from openpyxl import load_workbook

    from shared.tools.xlsx_render import markdown_to_xlsx

    md = (
        "# Risk Register\n\n| Risk | Impact |\n|---|---|\n| A | High |\n\n"
        "## Owners\n\n| Name | Role |\n|---|---|\n| Ana | Admin |\n"
    )
    out = str(tmp_path / "t.xlsx")
    markdown_to_xlsx(md, out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Risk Register", "Owners"]
    assert wb["Risk Register"]["A1"].value == "Risk"


@pytest.mark.unit
def test_a_sheet_name_excel_would_reject_is_made_legal(tmp_path):
    """Excel refuses a name over 31 characters or containing []:*?/\\ — openpyxl raises
    on save, losing the whole workbook rather than one tab."""
    from openpyxl import load_workbook

    from shared.tools.xlsx_render import markdown_to_xlsx

    md = "# A/B:C*D?E[a very long heading well past the excel limit]\n\n| X |\n|---|\n| 1 |\n"
    out = str(tmp_path / "n.xlsx")
    markdown_to_xlsx(md, out)
    name = load_workbook(out).sheetnames[0]
    assert len(name) <= 31
    assert not (set(name) & set("[]:*?/\\"))


@pytest.mark.unit
def test_two_tables_under_the_same_heading_get_distinct_sheets(tmp_path):
    """Excel rejects duplicate sheet names outright."""
    from openpyxl import load_workbook

    from shared.tools.xlsx_render import markdown_to_xlsx

    md = (
        "# Data\n\n| A |\n|---|\n| 1 |\n\ntext between\n\n| B |\n|---|\n| 2 |\n"
    )
    out = str(tmp_path / "dup.xlsx")
    markdown_to_xlsx(md, out)
    names = load_workbook(out).sheetnames
    assert len(names) == len(set(names)) == 2


@pytest.mark.unit
def test_a_document_with_no_tables_still_opens(tmp_path):
    """An empty workbook looks corrupt; a sheet saying why does not."""
    from openpyxl import load_workbook

    from shared.tools.xlsx_render import markdown_to_xlsx

    out = str(tmp_path / "none.xlsx")
    markdown_to_xlsx("# Just prose\n\nNo tables here.", out)
    ws = load_workbook(out).active
    assert "No tables" in str(ws["A1"].value)


@pytest.mark.unit
def test_the_separator_row_is_not_data(tmp_path):
    """|---|:--:|---| is layout. Exporting it would put dashes in row 2 of every sheet."""
    from openpyxl import load_workbook

    from shared.tools.xlsx_render import markdown_to_xlsx

    out = str(tmp_path / "sep.xlsx")
    markdown_to_xlsx("| A | B |\n|:--|--:|\n| 1 | 2 |\n", out)
    ws = load_workbook(out).active
    assert ws["A2"].value == "1"


# -- a row without bytes must not be reported as "saved" -----------------------


@pytest.mark.unit
async def test_a_failed_blob_upload_is_reported_separately(tmp_path):
    """From a live run:

        WARNING: Artifact upload failed for run 67fbd232-... (document): HttpResponseError
        INFO:  register_generated_file: persisted sdlc-password-reset-design.pdf

    store_artifact DEGRADES on an upload failure — it writes the row with
    blob_url=None so the document is still listed and the run does not die. Correct,
    and it was invisible: the user saw "saved to the project's artifacts", found the
    row in the panel, and found nothing in Blob.
    """
    from shared.services import chat_artifacts as ca

    f = tmp_path / "doc.pdf"
    f.write_bytes(b"%PDF-1.4")
    a, b, c, d = _ctx(consented=False)
    with a, b, c, d:
        await ca.register_generated_file("doc.pdf", str(f), "u", stage="design")

    async def _row_only(filename, file_path, url, *, stage, consented=None):
        # what store_artifact leaves behind when the upload fails
        ca._LAST_UPLOAD_OK[filename] = False

    with patch.object(ca, "register_generated_file", _row_only):
        stored, failed, not_uploaded = await ca.save_pending_artifacts("s1")

    assert stored == ["doc.pdf"]      # the row DID get written
    assert failed == []               # and nothing raised
    assert not_uploaded == ["doc.pdf"]  # but the bytes are not in storage


@pytest.mark.unit
async def test_a_successful_upload_is_not_flagged(tmp_path):
    from shared.services import chat_artifacts as ca

    f = tmp_path / "ok.pdf"
    f.write_bytes(b"%PDF-1.4")
    a, b, c, d = _ctx(consented=False)
    with a, b, c, d:
        await ca.register_generated_file("ok.pdf", str(f), "u", stage="design")

    async def _uploaded(filename, file_path, url, *, stage, consented=None):
        ca._LAST_UPLOAD_OK[filename] = True

    with patch.object(ca, "register_generated_file", _uploaded):
        stored, failed, not_uploaded = await ca.save_pending_artifacts("s1")

    assert stored == ["ok.pdf"] and not_uploaded == []


@pytest.mark.unit
@pytest.mark.parametrize(
    "rel",
    [
        "agents_orchestrator/requirements_agent/agents/planning.py",
        "agents_orchestrator/design_architecture_agent/agents/architecture.py",
    ],
)
def test_both_agents_warn_when_the_file_was_not_uploaded(rel):
    """The tool's reply has to distinguish the two, because they need different
    actions: nothing for a real save, an administrator for a row without bytes.

    Read by PATH rather than by import: both agents ship an `agents/planning.py`, and
    under pytest the dotted name resolved to the wrong one — a latent shadowing hazard
    worth avoiding here rather than debugging in a test about something else.
    """
    import re

    src = (Path(__file__).resolve().parents[1] / rel).read_text(encoding="utf-8")
    # Join adjacent string literals before flattening: the message is written across
    # several of them, so a naive whitespace squeeze leaves `the " "project's`.
    flat = re.sub(r'"\s*\n\s*"', "", src)
    flat = re.sub(r"\s+", " ", flat)
    assert "could not be uploaded to the project's file storage" in flat
    assert "the artifact is recorded and listed, but the file itself is not stored" in flat
