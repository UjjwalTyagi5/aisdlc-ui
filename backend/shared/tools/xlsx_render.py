"""Render Markdown tables to an .xlsx workbook.

WHAT THIS IS FOR. Both agents produce documents full of tables — a requirements
traceability matrix, a risk register, an API contract summary, a stakeholder list — and
a table is the one thing a reader wants OUT of the document and into a spreadsheet to
sort, filter and share. A .docx of a table is a picture of data.

ONE SHEET PER TABLE, named from the nearest heading above it, because a document with
four tables produces four sheets and "Sheet1..Sheet4" tells the reader nothing.

DELIBERATELY NOT a general Markdown converter: prose, lists and code blocks are dropped.
A spreadsheet of paragraphs is not useful, and pretending to convert them would produce
a file that looks like it worked.
"""
from __future__ import annotations

import logging
import os
import re
from typing import List, Tuple

logger = logging.getLogger(__name__)

#: Excel's own limits, not ours. A sheet name over 31 chars or containing any of these
#: raises deep inside openpyxl on save, losing the whole workbook.
_ILLEGAL_SHEET_CHARS = r'[]:*?/\\'
_MAX_SHEET_NAME = 31


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = "".join(" " if c in _ILLEGAL_SHEET_CHARS else c for c in (name or "")).strip()
    cleaned = re.sub(r"\s+", " ", cleaned) or "Table"
    cleaned = cleaned[:_MAX_SHEET_NAME]
    # Excel also rejects duplicates, and a numeric suffix must not push it over 31.
    base, n = cleaned, 2
    while cleaned.lower() in used:
        suffix = f" ({n})"
        cleaned = base[: _MAX_SHEET_NAME - len(suffix)] + suffix
        n += 1
    used.add(cleaned.lower())
    return cleaned


def extract_tables(content: str) -> List[Tuple[str, List[List[str]]]]:
    """Every Markdown pipe table in `content`, as (title, rows).

    The title is the nearest preceding heading, which is what makes the sheet tabs
    readable. A table with no heading above it gets "Table".
    """
    tables: List[Tuple[str, List[List[str]]]] = []
    heading = ""
    rows: List[List[str]] = []
    title_for_current = ""

    for raw in (content or "").replace("\r\n", "\n").split("\n"):
        line = raw.strip()
        h = re.match(r"^#{1,6}\s+(.*)$", line)
        if h:
            heading = h.group(1).strip()

        if line.startswith("|") and line.endswith("|") and len(line) > 1:
            cells = [c.strip() for c in line.strip("|").split("|")]
            # The |---|:--:|---| separator is layout, not a row.
            if cells and all(set(c) <= set("-: ") for c in cells if c):
                continue
            if not rows:
                title_for_current = heading
            rows.append(cells)
            continue

        if rows:
            tables.append((title_for_current, rows))
            rows = []
    if rows:
        tables.append((title_for_current, rows))
    return tables


def markdown_to_xlsx(content: str, output_path: str, sheet_title: str = "") -> str:
    """Write every Markdown table in `content` to `output_path`. Returns the path.

    A document with no tables still produces a workbook — with one sheet explaining
    that none were found, rather than an empty file that looks corrupt.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    tables = extract_tables(content)
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEF1F5")
    wrap = Alignment(vertical="top", wrap_text=True)
    used: set[str] = set()

    if not tables:
        ws = wb.create_sheet(_safe_sheet_name(sheet_title or "Document", used))
        ws["A1"] = "No tables were found in this document."
        ws["A2"] = "Markdown tables (| col | col |) are exported one sheet per table."
        ws.column_dimensions["A"].width = 70
        os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
        wb.save(output_path)
        return output_path

    for idx, (title, rows) in enumerate(tables, start=1):
        ws = wb.create_sheet(_safe_sheet_name(title or sheet_title or f"Table {idx}", used))
        for r, row in enumerate(rows, start=1):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.alignment = wrap
                if r == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        # Width from the longest cell, capped: one long sentence should not make a
        # column wider than the screen.
        for c in range(1, max(len(r) for r in rows) + 1):
            longest = max((len(str(r[c - 1])) for r in rows if len(r) >= c), default=10)
            ws.column_dimensions[get_column_letter(c)].width = min(max(longest + 2, 10), 60)
        ws.freeze_panes = "A2"

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path
